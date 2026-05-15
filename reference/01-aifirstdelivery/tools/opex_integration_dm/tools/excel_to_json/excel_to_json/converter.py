"""Excel-to-JSON conversion logic.

Reads each sheet of the analyst mapping workbook and emits a per-entity JSON config
shape conformant to /config/schemas/entity-mapping.schema.json.

What this tool does NOT do (and why):

- Does not invent `entitySetName` or `navigationProperty` for lookup transforms.
  These come from Dataverse metadata, not the workbook, and require a human pass.
  The tool leaves them as empty strings so `validate-config` flags them clearly.
- Does not parse the special trailing rows describing N:N relationships and post-load
  actions (Site sheet R19+, Contact R24+). Those are prose-shaped and brittle to parse.
  The analyst either edits the JSON directly or completes the workbook to a stricter
  shape in a future revision.
- Does not populate the freight/shipping/payment-term map values — those come from
  separate sub-workbooks the analyst has not yet provided (P0.8 in implementation.md).
"""
from __future__ import annotations

import json
import re
from pathlib import Path
from typing import Any, Iterator

import openpyxl

# ---------------------------------------------------------------------------
# Column mapping — workbook header → internal key
# ---------------------------------------------------------------------------
COLUMN_HEADERS = {
    "sr_no":           ["Sr No", "S No"],
    "source_system":   ["Source System"],
    "source_entity":   ["Source Entity"],
    "source_field":    ["Source Field"],
    "source_type":     ["Source Data Type"],
    "source_maxlen":   ["Source Field MaxLength", "Source Field Max Length"],
    "transform_logic": ["Transformation Logic"],
    "target_system":   ["Target System"],
    "target_entity":   ["Target Entity (Dataverse)", "Target Entity"],
    "target_display":  ["Target Filed", "Target Field"],
    "target_schema":   ["Target Field", "Target Field Schema"],
    "target_type":     ["Target Field Type", "Target Type"],
    "target_maxlen":   ["Target Field MaxLength"],
    "mandatory":       ["Mandatory (Y/N)", "Mandatory"],
    "default_value":   ["Default Value"],
    "key_type":        ["Key Type (PK/Alt)", "Key Type"],
    "validation_rule": ["Validation Rule"],
    "custom_field":    ["Custom Field"],
    "comments":        ["Comments"],
}


# ---------------------------------------------------------------------------
# Transform-type inference — "Transformation Logic" prose → transform.type
# ---------------------------------------------------------------------------
_INFERENCE_RULES: list[tuple[re.Pattern[str], str]] = [
    (re.compile(r"\bstatecode\b",                            re.I), "state"),
    (re.compile(r"\bmap.*choice\b|\bchoice\b.*map\b",        re.I), "choice"),
    (re.compile(r"\bmap.*lookup\b|\blookup\b.*code\b",       re.I), "lookup"),
    (re.compile(r"need to resolve lookup",                   re.I), "lookup"),
    (re.compile(r"check.*column.*value.*set as.+y\b",        re.I), "conditional"),
    (re.compile(r"\bmap.*y\s*/\s*n\b|\bmap\s+y\s*/\s*n\s+field\b", re.I), "yesNo"),
    (re.compile(r"\butc\s+time\b|\butc\s+time\s*zo[no]e\b",  re.I), "dateTime"),
    (re.compile(r"\bcheck.*value.*available.*choice\b",      re.I), "choice"),
    (re.compile(r"\bmap.*term.*code.*choice\b",              re.I), "choice"),
    (re.compile(r"\bmap.*freight.*term\b|\bmap.*shipping\b", re.I), "choice"),
    (re.compile(r"^direct$",                                 re.I), "direct"),
]


def infer_transform_type(logic_text: str | None) -> str:
    """Return the best-guess transform.type from the prose in 'Transformation Logic'."""
    if not logic_text:
        return "direct"
    text = str(logic_text).strip()
    if not text:
        return "direct"
    for pattern, t in _INFERENCE_RULES:
        if pattern.search(text):
            return t
    # Default fallback when prose exists but matches nothing recognized.
    return "direct"


# ---------------------------------------------------------------------------
# Target-type normalization
# ---------------------------------------------------------------------------
_TARGET_TYPE_MAP = {
    "text": "text",
    "string": "text",
    "text (phone number)": "text",
    "text (email)": "text",
    "integer": "int",
    "int": "int",
    "decimal": "decimal",
    "number": "decimal",
    "datetime": "datetime",
    "date only": "dateonly",
    "date": "dateonly",
    "date/time": "datetime",
    "datetime (date only)": "dateonly",
    "yes/no": "bool",
    "yesno": "bool",
    "bool": "bool",
    "boolean": "bool",
    "choice": "choice",
    "optionset": "choice",
    "state": "state",
    "statecode": "state",
    "lookup": "lookup",
}


def normalize_target_type(raw: str | None) -> str:
    if not raw:
        return "text"
    key = str(raw).strip().lower()
    return _TARGET_TYPE_MAP.get(key, "text")


def is_yes(cell: Any) -> bool:
    if cell is None:
        return False
    return str(cell).strip().upper() in {"Y", "YES", "TRUE", "1"}


# ---------------------------------------------------------------------------
# Sheet → entity dict
# ---------------------------------------------------------------------------
def _header_row(ws) -> dict[str, int]:
    """Map internal key → 1-based column index using the first row of the sheet."""
    raw_headers: list[str] = []
    for c in ws[1]:
        v = c.value
        raw_headers.append("" if v is None else str(v).strip().replace("\n", " "))
    out: dict[str, int] = {}
    for key, candidates in COLUMN_HEADERS.items():
        for col_idx, hdr in enumerate(raw_headers, start=1):
            if any(_header_match(hdr, cand) for cand in candidates):
                if key not in out:
                    out[key] = col_idx
    return out


def _header_match(actual: str, expected: str) -> bool:
    """Header strings can have line breaks / extra spaces; compare normalized."""
    a = re.sub(r"\s+", " ", actual or "").strip().lower()
    e = re.sub(r"\s+", " ", expected or "").strip().lower()
    return a == e


def _iter_data_rows(ws, hdr: dict[str, int]) -> Iterator[dict[str, Any]]:
    """Yield one dict per non-empty data row."""
    src_field_col = hdr.get("source_field")
    tgt_schema_col = hdr.get("target_schema")
    for r in ws.iter_rows(min_row=2, values_only=True):
        # Sheet has been padded to 102 rows with all-empty cells; skip those.
        if not any(c is not None and str(c).strip() != "" for c in r):
            continue
        # A row qualifies as a "field row" only if it has a source field OR a target schema name.
        sf = r[src_field_col - 1] if src_field_col else None
        ts = r[tgt_schema_col - 1] if tgt_schema_col else None
        if not (sf or ts):
            continue
        yield {k: (r[col - 1] if col and col <= len(r) else None) for k, col in hdr.items()}


def _row_to_field(row: dict[str, Any]) -> dict[str, Any] | None:
    """Convert one workbook row to a fields[] entry. Returns None if the row is invalid."""
    source = (row.get("source_field") or "").strip() if row.get("source_field") else None
    target = (row.get("target_schema") or "").strip() if row.get("target_schema") else None
    if not source or not target:
        return None

    target_type = normalize_target_type(row.get("target_type"))
    transform_type = infer_transform_type(row.get("transform_logic"))

    field: dict[str, Any] = {
        "source": source,
        "target": target,
        "targetType": target_type,
    }

    if row.get("source_type"):
        st = str(row["source_type"]).strip().lower()
        if "varchar" in st or "string" in st or "text" in st:
            field["sourceType"] = "text"
        elif "int" in st or "number" in st:
            field["sourceType"] = "int"
        elif "date" in st or "time" in st:
            field["sourceType"] = "datetime"
        elif "y / n" in st or "yes" in st or "bool" in st:
            field["sourceType"] = "bool"

    if row.get("source_maxlen") and str(row["source_maxlen"]).strip().isdigit():
        field["maxLength"] = int(str(row["source_maxlen"]).strip())
    if row.get("target_maxlen") and str(row["target_maxlen"]).strip().isdigit():
        field["maxLength"] = int(str(row["target_maxlen"]).strip())

    if is_yes(row.get("mandatory")):
        field["mandatory"] = True
        validations = ["notEmpty"]
        if "maxLength" in field:
            validations.append(f"maxLength:{field['maxLength']}")
        field["validations"] = validations

    if row.get("default_value") not in (None, ""):
        field["defaultValue"] = row["default_value"]

    # ----- transform block construction -----
    transform: dict[str, Any] = {"type": transform_type}
    if transform_type == "direct":
        pass
    elif transform_type == "state":
        transform["forward"] = {"map": {"A": 0, "I": 1}, "default": 0}
    elif transform_type == "choice":
        transform["forward"] = {"map": {}, "default": None, "onMissing": "fail"}
    elif transform_type == "lookup":
        # Leave navigationProperty/entitySetName empty for analyst to complete.
        transform["forward"] = {
            "referenceTable": "",
            "entitySetName": "",
            "navigationProperty": "",
            "refField": "",
            "primaryIdField": "",
            "matchPolicy": "exact",
            "createIfMissing": False,
            "onMissing": "error",
        }
    elif transform_type == "yesNo":
        transform["forward"] = {
            "trueValues":  ["Y", "Yes", "YES", "1", "true"],
            "falseValues": ["N", "No", "NO", "0", "false"],
        }
    elif transform_type == "dateTime":
        transform["forward"] = {
            "sourceTz": "UTC",
            "targetTz": "UTC",
            "inputFormat":  "yyyy-MM-dd HH:mm:ss",
            "outputFormat": "yyyy-MM-ddTHH:mm:ssZ",
        }
    elif transform_type == "conditional":
        transform["forward"] = {"rules": [], "default": None}

    field["transform"] = transform
    return field


def _sheet_target_entity(rows: list[dict[str, Any]]) -> str:
    """Pick the most common target entity logical name across all rows."""
    counts: dict[str, int] = {}
    for r in rows:
        v = (r.get("target_entity") or "").strip()
        if v:
            counts[v] = counts.get(v, 0) + 1
    if not counts:
        return ""
    # Most common; normalize a few common display names.
    name = max(counts, key=counts.get)
    return {
        "Account":              "account",
        "Contact":              "contact",
        "Functional Location":  "msdyn_functionallocation",
    }.get(name, name.lower().replace(" ", "_"))


def _sheet_alternate_keys(rows: list[dict[str, Any]]) -> list[str]:
    """Pull every field whose Key Type column says Alt Key or PK Alt."""
    keys: list[str] = []
    for r in rows:
        kt = (r.get("key_type") or "").strip().upper()
        ts = (r.get("target_schema") or "").strip()
        if ts and "ALT" in kt:
            keys.append(ts)
    return keys


def convert_sheet(ws, sheet_name: str | None = None) -> dict[str, Any]:
    """Read one workbook sheet and return its entity-mapping JSON shape."""
    sheet_name = sheet_name or ws.title
    entity_key = sheet_name.strip().lower()
    hdr = _header_row(ws)
    raw_rows = list(_iter_data_rows(ws, hdr))

    fields: list[dict[str, Any]] = []
    for r in raw_rows:
        f = _row_to_field(r)
        if f:
            fields.append(f)

    target_entity = _sheet_target_entity(raw_rows)
    alt_keys = _sheet_alternate_keys(raw_rows) or [(fields[0]["target"] if fields else "id")]

    return {
        "$schemaVersion": "1",
        "entityKey": entity_key,
        "direction": "inbound",
        "source": {
            "system": "SFTP",
            "format": "csv",
            "sftpLinkedService": "ls_sftp_fusion",
            "delimiter": ",",
            "encoding": "utf-8",
            "headerRow": 1,
            "inbound": {
                "pathPattern": f"/exports/{entity_key}_*.csv",
                "worksheet": None,
                "archiveTo": f"/archive/{entity_key}/{{yyyy}}/{{MM}}/{{dd}}/",
                "onMultipleFiles": "concat",
            },
        },
        "target": {
            "system": "Dataverse",
            "entityLogicalName": target_entity,
            "alternateKey": alt_keys,
            "inbound": {
                "writeMode": "upsert",
                "bypassPlugins": True,
            },
        },
        "dedup": {
            "enabled": True,
            "checksumColumns": [raw_rows[0].get("source_field")] if raw_rows else [],
            "onDuplicate": "keepLatest",
        },
        "dateTime": {"defaultTimeZone": "UTC", "format": "yyyy-MM-dd HH:mm:ss"},
        "lookups": {"createMissingDefault": False},
        "fields": fields,
        "relationships": [],
        "postLoadActions": [],
    }


def convert_workbook(workbook_path: Path, out_dir: Path) -> list[Path]:
    """Process every sheet of the workbook, emit one JSON per sheet."""
    wb = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)
    out_dir.mkdir(parents=True, exist_ok=True)
    written: list[Path] = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        doc = convert_sheet(ws, sheet_name)
        ek = doc["entityKey"]
        path = out_dir / f"{ek}.json"
        with path.open("w", encoding="utf-8") as fh:
            json.dump(doc, fh, indent=2, ensure_ascii=False)
            fh.write("\n")
        written.append(path)
    return written
